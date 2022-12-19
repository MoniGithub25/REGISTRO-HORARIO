import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl
import pyqrcode
import png
from pyqrcode import QRCode


cap = cv2.VideoCapture(0)

#VARIABLES
mañana = []
tarde = []
noche = []

#HORARIO
def infhora():
	#Informacion
	inf = datetime.now()
	#Extraemos fecha
	fecha = inf.strftime('%Y:%m:%d')
	#Extraemos hora
	hora = inf.strftime('%H:%M:%S')

	return hora, fecha


#PRINCIPAL
while True:
	#Leemos los frames
	ret, frame = cap.read()

	#INTERFAZ
	#Texto
	cv2.putText(frame, 'Ponga el codigo QR', (160,80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)
	#Ubicamos el rectangulo en las zonas extraidas
	cv2.rectangle(frame, (170,100), (470,400), (0,255,0), 2)

	#Extraemos hora y fecha
	hora, fecha = infhora()
	diasem = datetime.today().weekday()

	print(diasem)
	#AÑO MES DIA
	a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
	#Hora Minuto Segundo
	h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

	#Creamos archivo
	nomar= str(a) + '-' + str(me) + '-'+ str(d)
	texth = str(h) + ':' + str(m) + ':' + str(s)
	print(nomar)
	print(texth)
	#ARCHIVO EXCEL
	wb = xl.Workbook()

	#Leemos codigo QR
	for codes in decode(frame):

		#INFROMACION
		#Decodificamos
		info = codes.data.decode('utf-8')

		#Tipo de persona Letra
		tipo = info[0:2]
		tipo = int(tipo)
		letr = chr(tipo)


		#Numero
		num = info[2:]

		#Extraemos coordenadas
		pts = np.array([codes.polygon], np.int32)
		xi, yi = codes.rect.left, codes.rect.top


		#Redimensionamos
		pts = pts.reshape((-1,1,2))


		#ID completo
		codigo = letr + num


		#DIAS DE LA SEMANA
		if 4>= diasem >=0:

			#DIVIDIMOS LAS HORAS DEL DIA
			#MAÑANA
			if 9>= h >=7:
				cv2.polylines(frame, [pts], True, (255,255,0), 5)
				#Guardamos ID
				if codigo not in mañana:
					#Agregamos ID
					pos = len(mañana)
					mañana.append(codigo)


					#Guardamos QR
					hojam = wb.create_sheet("Llegada")
					datos = hojam.append(mañana)
					wb.save(nomar + '.xlsx')

					#Dibujamos
					cv2.putText(frame, letr + '0' + str(num), (xi-15, yi-15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,55,0),2)
				
				#Avisamos
				elif codigo in mañana:
					cv2.putText(frame, 'El ID '+str(codigo),(xi-65,yi-45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0),2)
					cv2.putText(frame, 'Fue registrado',(xi-65,yi-15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0),2)

			#TARDE
			if 15>= h >=14:
				cv2.polylines(frame, [pts], True, (255,255,0), 5)
				#Guardamos ID
				if codigo not in tarde:
					#Agregamos ID
					tarde.append(codigo)


					#Guardamos QR
					hojat = wb.create_sheet("Salida")
					datos = hojat.append(tarde)
					wb.save(nomar + '.xlsx')

					#Dibujamos
					cv2.putText(frame, letr + '0' + str(num), (xi-15, yi-15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,55,0),2)

				#Avisamos
				elif codigo in tarde:
					cv2.putText(frame, 'El ID '+str(codigo),(xi-65,yi-45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0),2)
					cv2.putText(frame, 'Fue registrado',(xi-65,yi-15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0),2)

	# Mostramos FPS
	cv2.imshow("LECTOR DE QR", frame)
	# Leemos teclado
	t = cv2.waitKey(5)
	if t == 27:
		break

cv2.destroyAllWindows()
cap.release()				



